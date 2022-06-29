import statistics
import openpyxl
import math
from openpyxl import load_workbook
wb = load_workbook(filename = 'normal_table2.xlsx', read_only=True)
wh = load_workbook(filename = 'student_table.xlsx', read_only=True)

print ('Выберите случай для нахождения доверительного интервала: 1) Дана выборка и дисперсия; 2) Дан объем выборки, среднее значение и стандартное отклонение; 3) Дан объём выборки, выборочное среднее, выборочная дисперсия; 4) Дана заданная точность, дисперсия, уровень доверия и необходимо найти объем выборки;')
variant = int(input())

if variant is 1: 
    raw = input('Введите последовательность чисел через пробел: ')
    n = [int(i) for i in raw.split(' ') if i.isdigit()]
    print('Введён массив ', n)
    print('Кол-во элментов выборки: ', len(n))
    N = float(len(n))
    print('Введите дисперсию σ^2: ')
    o = float(input())
    print('Введите доверительный интервал в инетвал 0<B<1: ')
    a = 1 - float(input())
    middle = statistics.mean(n) 
    print('Cреднее значение: ', middle)
    B = 1 - (a/2);
    print('бета: ', B)
    
    proverka = int(0)
    
    for row in wb['Лист1'].iter_rows(): #пробежим по таблице
        for cell in row:
            if float(cell.value) == round(B, 4): #нашли значение
                print('квантиль приблизительно Zo =',wb['Лист1'].cell(row=1, column=cell.column).value+wb['Лист1'].cell(row=cell.row, column=1).value)
                proverka = 1
                break
            
    proverka2 = int(0)
    
    if proverka == 0:
        for row in wb['Лист1'].iter_rows(): #пробежим по таблице
            if proverka2==0: 
                for cell in row:
                   if round(float(cell.value), 3) == round(B, 3):#нашли значение 
                      Zo = float(float(wb['Лист1'].cell(row=1, column=cell.column).value)+float(wb['Лист1'].cell(row=cell.row, column=1).value))
                      print('квантиль приблизительно Zo =', Zo)
                      proverka2=1;
                      break
    if proverka == 0:
        for row in wb['Лист1'].iter_rows(): #пробежим по таблице
            if proverka2==0: 
                for cell in row:
                   if round(float(cell.value), 2) == round(B, 2):#нашли значение 
                      Zo = float(float(wb['Лист1'].cell(row=1, column=cell.column).value)+float(wb['Лист1'].cell(row=cell.row, column=1).value))
                      print('квантиль приблизительно Zo =', Zo)
                      proverka2=1;
                      break
               
    delta = round(float(math.sqrt(o/N)*Zo), 2)
    print('точность =', delta)
    print('Доверительный инвервал имеет вид (', middle - delta, ' ; ', middle + delta, ') ')
    
elif variant is 2: 
    print('Введите объем выборки n= ')
    n = float(input())
    print('Введите среднее значение Х*= ')
    X = float(input())
    print('Введите cтандартное отклонение о= ')
    o = float(input())
    print('Введите доверительный интервал в инетвале 0<B<1: ')
    a = 1 - float(input())
    print('уровень значимости а: ', a)
    B = 1 - (a/2);
    print('бета: ', B)
    
    proverka = int(0)
    
    for row in wb['Лист1'].iter_rows(): #пробежим по таблице
        for cell in row:
            if float(cell.value) == round(B, 4): #нашли значение
                print('квантиль приблизительно Zo =',wb['Лист1'].cell(row=1, column=cell.column).value+wb['Лист1'].cell(row=cell.row, column=1).value)
                proverka = 1
                break
            
    proverka2 = int(0)
    
    if proverka == 0:
        for row in wb['Лист1'].iter_rows(): #пробежим по таблице
            if proverka2==0: 
                for cell in row:
                   if round(float(cell.value), 3) == round(B, 3):#нашли значение 
                      Zo = float(float(wb['Лист1'].cell(row=1, column=cell.column).value)+float(wb['Лист1'].cell(row=cell.row, column=1).value))
                      print('квантиль приблизительно Zo =', Zo)
                      proverka2=1;
                      break
    if proverka == 0:
        for row in wb['Лист1'].iter_rows(): #пробежим по таблице
            if proverka2==0: 
                for cell in row:
                   if round(float(cell.value), 2) == round(B, 2):#нашли значение 
                      Zo = float(float(wb['Лист1'].cell(row=1, column=cell.column).value)+float(wb['Лист1'].cell(row=cell.row, column=1).value))
                      print('квантиль приблизительно Zo =', Zo)
                      proverka2=1;
                      break
    delta = round(float(math.sqrt(o/n)*Zo), 3)
    print('точность =', delta)
    print('Доверительный инвервал имеет вид (', X - delta, ' ; ', X + delta, ') ')

elif variant is 3: 
    print('Введите малый объём выборки(n<30): ') 
    n = float(input())
    if n > 31:
            print('ошибка данных')           
    print('Введите выборочное среднее значение x¯= ')
    x = float(input())
    print('Введите выборочную дисперсию s^2= ')
    s = float(input())
    print('Введите доверительный интервал в инетвале 0<B<1: ')
    a = 1 - float(input())
    B = 1 - (a/2);
    print('бета: ', B)
    step = int(n-1)
    print('Кол-во степеней свободы n-1: ', step)
    
    #находим квантиль t(a) по твблице Стьюодента
    
    stroka = int() #строка в которой квантиль
    proverka = int(0)
    for row in wh['Лист2'].iter_rows(): #пробежим по cтепеням свободы
        for cell in row:     
           if int(cell.value) == step: #нашли значение
               stroka = cell.row
               print('//строка = ', stroka) #здесь возьмем координату по строкам
               proverka = 1
               stroka = cell.row
               break
    if proverka == 0:
        for row in wh['Лист2'].iter_rows(): 
           for cell in row:     
               if int(cell.value) == step-1: #нашли значение
                   stroka = cell.row
                   print('//строка = ', stroka) #здесь возьмем координату по строкам
                   proverka = 1
                   break
               
    stolbec = float() #столбец в которой квантиль
    proverka2 = int(0)
    for row in wh['Лист3'].iter_rows(): #пробежим по вероятностям
        for cell in row:     
            if float(cell.value) == B:
                stolbec = cell.column
                print('//столбец = ', stolbec)
                proverka2 = 1;
                break
    if proverka2 == 0:
          for row in wh['Лист3'].iter_rows(): #пробежим по вероятностям
              for cell in row:     
                  if round(float(cell.value),2) == round(B, 2):
                     stolbec = cell.column
                     print('//столбец = ', stolbec)
                     proverka2 = 1
                     break
    if proverka2 == 0:
        print('данное значение по вероятности B не найдено, присвоенно близкое значение')
        print('//столбец = 3')
        stolbec = 3;
    t = float(wh['Лист1'].cell(row=stroka, column=stolbec).value)
    print('Квантиль t(a) по таблице Стьюдента= ', str(t))
    tochn = round(float(math.sqrt(s/n)*t), 3)
    print('Точность = (s/sqr(n))*t(a) =', str(tochn))
    print('Доверительный инвервал имеет вид (', round(x - tochn, 2), ' ; ', round(x + tochn, 2), ') ')

elif variant is 4: 
    print('Укажите точность delta: ')
    delta = float(input())
    print('Укажите дисперсию в квадрате o^2: ')
    o = float(input())
    print('Введите уровень доверия в инетвале 0<B<1: ')
    B = float(input())
    
    proverka = int(0)
    
    for row in wb['Лист1'].iter_rows(): #пробежим по таблице
        for cell in row:
            if float(cell.value) == round(B, 4): #нашли значение
                print('квантиль приблизительно Zo =',wb['Лист1'].cell(row=1, column=cell.column).value+wb['Лист1'].cell(row=cell.row, column=1).value)
                proverka = 1
                break
            
    proverka2 = int(0)
    
    if proverka == 0:
        for row in wb['Лист1'].iter_rows(): #пробежим по таблице
            if proverka2==0: 
                for cell in row:
                   if round(float(cell.value), 3) == round(B, 3):#нашли значение 
                      Zo = float(float(wb['Лист1'].cell(row=1, column=cell.column).value)+float(wb['Лист1'].cell(row=cell.row, column=1).value))
                      print('квантиль приблизительно Zo =', round(Zo, 3))
                      proverka2=1;
                      break
    if proverka == 0:
        for row in wb['Лист1'].iter_rows(): #пробежим по таблице
            if proverka2==0: 
                for cell in row:
                   if round(float(cell.value), 2) == round(B, 2):#нашли значение 
                      Zo = float(float(wb['Лист1'].cell(row=1, column=cell.column).value)+float(wb['Лист1'].cell(row=cell.row, column=1).value))
                      print('квантиль приблизительно Zo =', round(Zo, 3))
                      proverka2=1;
                      break
    n = o*math.pow((Zo/delta), 2)
    if n < 1:
        n = 1
    print('Минимальный объем выборки =', int(n))
else: 
    print('Неверный номер')
