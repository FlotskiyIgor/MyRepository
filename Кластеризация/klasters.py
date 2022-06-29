import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from sklearn.datasets import make_blobs
from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
X_xl = load_workbook(filename = 'X_coor.xlsx', read_only=True)
Y_xl = load_workbook(filename = 'Y_coor.xlsx', read_only=True)
DIM = 2 #кол-во измерений

print ('Выберите вариант: 1 - сгенерировать данные, 2 - загрузить данные')
choose = int(input())
if (choose == 1):
    print ('Введите оптимальное кол-во точек')
    N = int(input()) #кол-во точек
    
    #Генерируем координаты
    x = np.random.randn(N, DIM)  
    y = np.zeros(N)

elif (choose == 2):
    print('Загрузка данных')
    X_xl = pd.read_excel('X_coor.xlsx')
    Y_xl = pd.read_excel('Y_coor.xlsx')
    X_train = X_xl.to_numpy()
    N=100
    y = np.zeros(N)
    Y_train = Y_xl.to_numpy()
    X_train = X_train.astype('float64')
    Y_train = Y_train.astype('float64')
    x = X_train
else:
    print('Ошибка')
    input()

crit = []
F,d = make_blobs(N,n_features=2, centers=2)
for k in range(2, 8):
    kmeans = KMeans(n_clusters = k)
    kmeans.fit(F)
    crit.append(kmeans.inertia_)  
plt.plot(range(2,8), crit) #Вывод графика (Метод локтя)
plt.title(r'Метод локтя', fontsize = 20)
plt.show()

print ('Введите оптимальное кол-во кластеров на переломе графика')
num_cluster = int(input()) #кол-во кластеров

print ('Введите кол-во итераций')
iterations = int(input()) #кол-во итераций

#отрисовка изначальных данных
for k in range(1):
    plt.scatter(x[y==k,0], x[y==k,1])
plt.title(r'Отрисовка изначальных данных', fontsize = 20)
plt.show()

for t in range(iterations): #цикл зависит от кол-ва итераций
    if t == 0:
        index_ = np.random.choice(range(N),num_cluster,replace=False)
        mean = x[index_] #инекст показывает принадлежность к кластеру 
    else:
        for k in range(num_cluster): #примерно находит среднее расстояние от центроида
            mean[k] = np.mean(x[y==k], axis=0) #находит центроиду
    for i in range(N):
        dist = np.sum((mean - x[i])**2, axis=1) #дистанция
        pred = np.argmin(dist) #принадлежность к кластеру, это номер кластера
        y[i] = pred
    for k in range(num_cluster): #отрисовка  для каждого кластера
        #Отрисовка центроид
        plt.scatter(mean[:, 0], mean[:, 1], c = 'black', s = 60, alpha = 0.9);
        #Отрисовка точек
        fig = plt.scatter(x[y==k,0], x[y==k,1])
    plt.show()
